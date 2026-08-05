#!/usr/bin/env python3
"""Import DroneEdge School District Research.xlsx into outreach contact CSVs.

Reads the wide research workbook (one sheet per county) and emits normalized
contact rows matching docs/sales/contact-collection.md.

Usage:
  python scripts/import_district_research.py
  python scripts/import_district_research.py --merge-candidates

Requires openpyxl (scripts/.venv or any env with openpyxl installed).
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit(
        "openpyxl is required. Try: python3 -m venv /tmp/xlsx-venv && "
        "/tmp/xlsx-venv/bin/pip install openpyxl && "
        "/tmp/xlsx-venv/bin/python scripts/import_district_research.py"
    ) from e

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "outreach" / "DroneEdge School District Research.xlsx"
OUT_RESEARCH = ROOT / "outreach" / "school-district-research-contacts.csv"
OUT_CANDIDATES = ROOT / "outreach" / "contact-candidates.csv"

FIELDNAMES = [
    "organization",
    "organization_type",
    "state",
    "county",
    "source_url",
    "source_title",
    "source_type",
    "contact_name",
    "title",
    "department",
    "email",
    "phone",
    "role_tags",
    "confidence",
    "review_status",
    "notes",
    "collected_at",
]

SHEET_META = {
    "Delco": ("PA", "Delaware"),
    "Montco": ("PA", "Montgomery"),
    "Bucks": ("PA", "Bucks"),
    "Chester": ("PA", "Chester"),
    "Philadelphia": ("PA", "Philadelphia"),
    "Glocester(NJ)": ("NJ", "Gloucester"),
    "Camden(NJ)": ("NJ", "Camden"),
    "Burlington(NJ)": ("NJ", "Burlington"),
    "New Castle(DE)": ("DE", "New Castle"),
}

SOURCE_TITLE = "DroneEdge School District Research"
SOURCE_URL = "outreach/DroneEdge School District Research.xlsx"
SOURCE_TYPE = "research_workbook"

INTERNAL_EMAIL_RE = re.compile(
    r"internal|messaging|request form|\?\?\?\?|none",
    re.I,
)
EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def extract_email(raw: str) -> tuple[str, str]:
    """Return (email, note). Note captures assistant / non-email caveats."""
    raw = clean(raw)
    if not raw:
        return "", ""
    if INTERNAL_EMAIL_RE.search(raw) and "@" not in raw:
        return "", f"email_unavailable: {raw}"
    match = EMAIL_RE.search(raw)
    if not match:
        if INTERNAL_EMAIL_RE.search(raw):
            return "", f"email_unavailable: {raw}"
        return "", f"email_unparsed: {raw}" if raw else ""
    email = match.group(0).lower()
    note = ""
    lower = raw.lower()
    if "assistant" in lower:
        note = "listed as assistant contact / inbox"
    elif "(" in raw and raw.strip() != email:
        # keep parenthetical context without the email itself
        extra = EMAIL_RE.sub("", raw).strip(" ()")
        if extra:
            note = extra
    return email, note


def clean_phone(raw: str) -> str:
    phone = clean(raw)
    if not phone or phone.lower() in {"none", "n/a"}:
        return ""
    # Drop leading junk like lone ')' from OCR-ish entries
    phone = phone.lstrip(")(").strip()
    return phone


def role_tags_for(title: str, is_superintendent: bool) -> str:
    t = title.lower()
    tags: list[str] = []
    if is_superintendent or "superintendent" in t:
        tags.append("leadership")
    if any(k in t for k in ("cte", "career and technical", "vocational", "technical college")):
        tags.append("cte")
    if any(k in t for k in ("stem", "steam", "science", "technology", "engineering", "math")):
        tags.append("stem")
    if any(k in t for k in ("curriculum", "teaching", "instruction", "academic", "learning")):
        tags.append("curriculum")
    if "grant" in t:
        tags.append("grants")
    if not tags:
        tags.append("curriculum" if title else "leadership")
    # stable unique order
    order = ["cte", "stem", "curriculum", "grants", "leadership"]
    return ";".join([x for x in order if x in tags])


def confidence_for(email: str, phone: str, notes: str) -> str:
    if email:
        return "high"
    if phone and "email_unavailable" not in notes:
        return "medium"
    return "low"


def org_type_for(district: str) -> str:
    d = district.lower()
    if any(k in d for k in ("intermediate unit", "technical college", "vocational", "institute of technology", "county technical")):
        return "service_agency"
    return "district"


def parse_contact_blocks(cells: list) -> list[dict]:
    """Parse superintendent + repeating curriculum blocks from a row."""
    contacts: list[dict] = []
    district = clean(cells[0]) if cells else ""
    if not district or district.lower().startswith("(note"):
        return []

    # Skip meta-only rows (e.g. CCIU pointer without people)
    padded = list(cells) + [None] * max(0, 16 - len(cells))

    # Superintendent block: name, email, phone
    super_name = clean(padded[1])
    super_email, super_note = extract_email(clean(padded[2]))
    super_phone = clean_phone(clean(padded[3]))
    if super_name:
        contacts.append(
            {
                "contact_name": super_name,
                "title": "Superintendent",
                "email": super_email,
                "phone": super_phone,
                "is_superintendent": True,
                "extra_note": super_note,
            }
        )

    # Curriculum blocks start at index 4: name, title, email, phone
    i = 4
    while i < len(padded):
        name = clean(padded[i]) if i < len(padded) else ""
        title = clean(padded[i + 1]) if i + 1 < len(padded) else ""
        email_raw = clean(padded[i + 2]) if i + 2 < len(padded) else ""
        phone_raw = clean(padded[i + 3]) if i + 3 < len(padded) else ""

        # Stop when the remaining block is empty
        if not any([name, title, email_raw, phone_raw]):
            i += 4
            continue

        # Misaligned rows: phone landed in email slot, or email-only without name
        email, email_note = extract_email(email_raw)
        phone = clean_phone(phone_raw)
        if not email and email_raw and re.search(r"\d", email_raw) and "@" not in email_raw:
            # e.g. Bensalem: phone put where email should be
            if not phone:
                phone = clean_phone(email_raw)
            email_note = (email_note + "; " if email_note else "") + "phone_was_in_email_column"

        # Name missing but email present — still keep as office contact
        if not name and not email and not phone:
            i += 4
            continue
        if not name and (email or title):
            name = title or email.split("@")[0]
            title = title or "Curriculum contact"

        contacts.append(
            {
                "contact_name": name,
                "title": title or "Curriculum / Instruction",
                "email": email,
                "phone": phone,
                "is_superintendent": False,
                "extra_note": email_note,
            }
        )
        i += 4

    return contacts


def sheet_rows(ws) -> list[list]:
    out = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        while vals and vals[-1] is None:
            vals.pop()
        out.append(vals)
    return out


def import_workbook(path: Path, collected_at: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    records: list[dict] = []
    for sheet_name, (state, county) in SHEET_META.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows = sheet_rows(wb[sheet_name])
        for row in rows:
            if not row:
                continue
            first = clean(row[0]).lower()
            if not first or first == "school district" or first.startswith("(note"):
                continue
            district = clean(row[0])
            for person in parse_contact_blocks(row):
                notes_parts = [
                    f"county={county}",
                    f"sheet={sheet_name}",
                ]
                if person["extra_note"]:
                    notes_parts.append(person["extra_note"])
                notes = "; ".join(notes_parts)
                email = person["email"]
                phone = person["phone"]
                records.append(
                    {
                        "organization": district,
                        "organization_type": org_type_for(district),
                        "state": state,
                        "county": county,
                        "source_url": SOURCE_URL,
                        "source_title": SOURCE_TITLE,
                        "source_type": SOURCE_TYPE,
                        "contact_name": person["contact_name"],
                        "title": person["title"],
                        "department": "Central Office" if person["is_superintendent"] else "Curriculum / Instruction",
                        "email": email,
                        "phone": phone,
                        "role_tags": role_tags_for(person["title"], person["is_superintendent"]),
                        "confidence": confidence_for(email, phone, notes),
                        "review_status": "needs_review",
                        "notes": notes,
                        "collected_at": collected_at,
                    }
                )
    return records


def dedupe_key(row: dict) -> str:
    email = (row.get("email") or "").strip().lower()
    if email:
        return f"email:{email}"
    org = (row.get("organization") or "").strip().lower()
    name = (row.get("contact_name") or "").strip().lower()
    title = (row.get("title") or "").strip().lower()
    return f"name:{org}|{name}|{title}"


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_candidates(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def merge_into_candidates(research_rows: list[dict], candidates_path: Path) -> tuple[int, int]:
    existing = read_candidates(candidates_path)
    seen = {dedupe_key(r) for r in existing}
    # Candidates schema may omit county — drop it when merging
    candidate_fields = [
        "organization",
        "organization_type",
        "state",
        "source_url",
        "source_title",
        "source_type",
        "contact_name",
        "title",
        "department",
        "email",
        "phone",
        "role_tags",
        "confidence",
        "review_status",
        "notes",
        "collected_at",
    ]
    added = 0
    skipped = 0
    for row in research_rows:
        key = dedupe_key(row)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        existing.append({k: row.get(k, "") for k in candidate_fields})
        added += 1
    write_csv(candidates_path, existing, candidate_fields)
    return added, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=OUT_RESEARCH)
    parser.add_argument(
        "--merge-candidates",
        action="store_true",
        help=f"Also append new rows into {OUT_CANDIDATES.name} (deduped by email/name).",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Workbook not found: {args.xlsx}")

    collected_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    rows = import_workbook(args.xlsx, collected_at)
    write_csv(args.out, rows, FIELDNAMES)

    by_state: dict[str, int] = {}
    with_email = 0
    for r in rows:
        by_state[r["state"]] = by_state.get(r["state"], 0) + 1
        if r["email"]:
            with_email += 1

    print(f"Wrote {len(rows)} contacts → {args.out.relative_to(ROOT)}")
    print(f"  with email: {with_email}")
    print(f"  by state: {by_state}")

    if args.merge_candidates:
        added, skipped = merge_into_candidates(rows, OUT_CANDIDATES)
        print(f"Merged into {OUT_CANDIDATES.relative_to(ROOT)}: +{added} new, {skipped} duplicates skipped")


if __name__ == "__main__":
    main()
